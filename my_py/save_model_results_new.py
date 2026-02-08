from openpyxl import load_workbook
import os
import sys

def save_model_results_new(
    model_params,
    rf_score,
    rf_submission,
    performance_results,
    train_ratio,
    step_time,
    min_sleep_duration=12*30,
    output_file='sleep_data.xlsx'
):
    """
    保存模型结果到Excel文件
    
    参数:
        model_params: 模型超参数字典
        rf_score: 模型得分
        rf_submission: 预测结果
        performance_results: 性能结果字典
        train_ratio: 训练集比例
        step_time: 步长时间
        min_sleep_duration: 最小睡眠持续时间
        output_file: 输出Excel文件名
    
    返回:
        int: 保存的行号
    """
    try:
        # 从model_params中提取参数
        n_estimators = model_params.get('n_estimators', 100)
        criterion = model_params.get('criterion', 'gini')
        max_depth = model_params.get('max_depth', None)
        min_samples_split = model_params.get('min_samples_split', 2)
        min_samples_leaf = model_params.get('min_samples_leaf', 1)
        min_weight_fraction_leaf = model_params.get('min_weight_fraction_leaf', 0.0)
        max_features = model_params.get('max_features', 'sqrt')
        max_leaf_nodes = model_params.get('max_leaf_nodes', None)
        min_impurity_decrease = model_params.get('min_impurity_decrease', 0.0)
        bootstrap = model_params.get('bootstrap', True)
        oob_score = model_params.get('oob_score', False)
        n_jobs = model_params.get('n_jobs', None)
        random_state = model_params.get('random_state', None)
        verbose = model_params.get('verbose', 0)
        warm_start = model_params.get('warm_start', False)
        class_weight = model_params.get('class_weight', None)
        ccp_alpha = model_params.get('ccp_alpha', 0.0)
        max_samples = model_params.get('max_samples', None)
        monotonic_cst = model_params.get('monotonic_cst', None)
        
        # 构建第一列数据
        first_column_data = [
            n_estimators,
            criterion,
            max_depth,
            min_samples_split,
            min_samples_leaf,
            min_weight_fraction_leaf,
            max_features,
            max_leaf_nodes,
            min_impurity_decrease,
            bootstrap,
            oob_score,
            n_jobs,
            random_state,
            verbose,
            warm_start,
            class_weight,
            ccp_alpha,
            max_samples,
            monotonic_cst,
            train_ratio,
            step_time,
            min_sleep_duration
        ]
        
        # 将第一列数据转换为字符串，用'/'分隔
        first_column_str = '/'.join(str(item) if item is not None else 'None' for item in first_column_data)
        
        # 从performance_results中提取数据
        model_size = '0 MB'
        total_val_days = 0
        prediction_days = 0
        predicted_onset_count = 0
        predicted_wakeup_count = 0
        prediction_rate = 0.0
        onset_event_rate = 0.0
        wakeup_event_rate = 0.0
        
        # 初始化匹配时间统计数据
        onset_match_counts = ['0'] * 6
        onset_match_rates = ['0.0'] * 6
        onset_mean_diffs = ['0.0'] * 6
        onset_std_diffs = ['0.0'] * 6
        wakeup_match_counts = ['0'] * 6
        wakeup_match_rates = ['0.0'] * 6
        wakeup_mean_diffs = ['0.0'] * 6
        wakeup_std_diffs = ['0.0'] * 6
        total_match_counts = ['0'] * 6
        total_match_rates = ['0.0'] * 6
        total_mean_diffs = ['0.0'] * 6
        total_std_diffs = ['0.0'] * 6
        
        if performance_results is not None:
            try:
                # 提取模型信息
                if 'model_info' in performance_results:
                    model_info = performance_results['model_info']
                    model_size_mb = model_info.get('model_size_mb', 0)
                    model_size = f"{model_size_mb:.2f} MB"
                    predicted_onset_count = model_info.get('predicted_onset_count', 0)
                    predicted_wakeup_count = model_info.get('predicted_wakeup_count', 0)
                
                # 提取预测统计信息
                if 'prediction_stats' in performance_results:
                    prediction_stats = performance_results['prediction_stats']
                    prediction_rate = prediction_stats.get('prediction_rate', 0.0)
                    prediction_days = prediction_stats.get('total_predicted_days', 0)
                    onset_event_rate = prediction_stats.get('onset_event_rate', 0.0)
                    wakeup_event_rate = prediction_stats.get('wakeup_event_rate', 0.0)
                
                # 提取总验证天数
                total_val_days = performance_results.get('total_val_days', 0)
                
                # 提取匹配时间统计信息
                if 'match_time_stats' in performance_results:
                    match_time_stats = performance_results['match_time_stats']
                    # 按匹配时间排序
                    sorted_match_times = sorted(match_time_stats.keys())
                    
                    # 提取各种统计数据
                    for i, match_time in enumerate(sorted_match_times):
                        if i >= 6:  # 只取前6个时间点
                            break
                        
                        stats = match_time_stats[match_time]
                        
                        # Onset统计
                        onset_stats = stats.get('onset', {})
                        onset_match_counts[i] = str(onset_stats.get('matches', 0))
                        onset_match_rates[i] = f"{onset_stats.get('match_rate', 0.0):.4f}"
                        onset_mean_diffs[i] = f"{onset_stats.get('avg_time_diff', 0.0):.2f}"
                        onset_std_diffs[i] = f"{onset_stats.get('std_time_diff', 0.0):.2f}"
                        
                        # Wakeup统计
                        wakeup_stats = stats.get('wakeup', {})
                        wakeup_match_counts[i] = str(wakeup_stats.get('matches', 0))
                        wakeup_match_rates[i] = f"{wakeup_stats.get('match_rate', 0.0):.4f}"
                        wakeup_mean_diffs[i] = f"{wakeup_stats.get('avg_time_diff', 0.0):.2f}"
                        wakeup_std_diffs[i] = f"{wakeup_stats.get('std_time_diff', 0.0):.2f}"
                        
                        # 总统计
                        total_stats = stats.get('total', {})
                        total_match_counts[i] = str(total_stats.get('matches', 0))
                        total_match_rates[i] = f"{total_stats.get('match_rate', 0.0):.4f}"
                        total_mean_diffs[i] = f"{total_stats.get('avg_time_diff', 0.0):.2f}"
                        total_std_diffs[i] = f"{total_stats.get('std_time_diff', 0.0):.2f}"
            except (KeyError, AttributeError, TypeError) as e:
                print(f"警告: 从performance_results提取数据时出错: {e}")
                print("使用默认值继续保存数据")
        
        # 构建数据列表
        data = [
            first_column_str,
            model_size,
            f"{total_val_days}/{prediction_days}/{predicted_onset_count}/{predicted_wakeup_count}",
            f"{prediction_rate:.4f}/{onset_event_rate:.4f}/{wakeup_event_rate:.4f}",
            '/'.join(onset_match_counts),
            '/'.join(onset_match_rates),
            '/'.join(onset_mean_diffs),
            '/'.join(onset_std_diffs),
            '/'.join(wakeup_match_counts),
            '/'.join(wakeup_match_rates),
            '/'.join(wakeup_mean_diffs),
            '/'.join(wakeup_std_diffs),
            '/'.join(total_match_counts),
            '/'.join(total_match_rates),
            '/'.join(total_mean_diffs),
            '/'.join(total_std_diffs),
            rf_score
        ]
        
        file_path = f'/home/zhuangzhuohan/sleep_data/{output_file}'
        
        # 检查文件权限
        try:
            if os.path.exists(file_path):
                if not os.access(file_path, os.R_OK):
                    raise PermissionError(f"无法读取文件: {file_path}")
                if not os.access(file_path, os.W_OK):
                    raise PermissionError(f"无法写入文件: {file_path}")
        except PermissionError as e:
            print(f"错误: {e}")
            return None
        
        # 加载或创建Excel文件
        try:
            headers = [
                '模型参数',
                '模型内存占用',
                '总天数/预测天数/预测onset事件数/预测wakeup事件数',
                '预测率/预测onset事件率/预测wakeup事件率',
                'onset匹配数(5/10/15/30/60/120)分钟',
                'onset匹配率(5/10/15/30/60/120)分钟',
                'onset平均差(秒)(5/10/15/30/60/120)分钟',
                'Onset标准差(秒)(5/10/15/30/60/120)分钟',
                'Wakeup匹配数(5/10/15/30/60/120)分钟',
                'Wakeup匹配率(5/10/15/30/60/120)分钟',
                'Wakeup平均差(秒)(5/10/15/30/60/120)分钟',
                'Wakeup标准差(秒)(5/10/15/30/60/120)分钟',
                '总和匹配数(5/10/15/30/60/120)分钟',
                '总和匹配率(5/10/15/30/60/120)分钟',
                '总和平均差(秒)(5/10/15/30/60/120)分钟',
                '总和标准差(秒)(5/10/15/30/60/120)分钟',
                '模型得分'
            ]
            
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                # 查找包含正确表头的工作表
                target_sheet = None
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # 检查第一行是否包含预期的表头
                    if ws.max_row >= 1:
                        first_row_values = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
                        if first_row_values == headers:
                            target_sheet = ws
                            break
                
                # 如果找到目标工作表，使用它
                if target_sheet:
                    ws = target_sheet
                    # 直接使用max_row + 1，确保新数据总是添加到文件末尾
                    # 这样即使中间有空行，新数据也会添加到文件末尾
                    next_row = ws.max_row + 1
                    wb.active = ws
                else:
                    # 如果没有找到，创建一个新的工作表
                    ws = wb.create_sheet(title='Sleep Data')
                    for col_idx, header in enumerate(headers, start=1):
                        ws.cell(row=1, column=col_idx, value=header)
                    next_row = 2
                    wb.active = ws
            else:
                # 创建新的Excel文件
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'Sleep Data'
                # 添加表头
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
                next_row = 2
                wb.active = ws
        except Exception as e:
            print(f"错误: 加载或创建Excel文件失败: {e}")
 
        # 写入数据
        try:
            for col_idx, value in enumerate(data, start=1):
                ws.cell(row=next_row, column=col_idx, value=value)
        except Exception as e:
            print(f"错误: 写入数据到Excel失败: {e}")
            return None
        
        # 保存文件
        try:
            wb.save(file_path)
            print(f"数据已保存到：{file_path}")
            print(f"行号：{next_row}")
            return next_row
        except Exception as e:
            print(f"错误: 保存Excel文件失败: {e}")
            return None
    
    except Exception as e:
        print(f"错误: 保存模型结果时发生未知错误: {e}")
        import traceback
        traceback.print_exc()
        return None