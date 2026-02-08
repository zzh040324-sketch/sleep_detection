from openpyxl import load_workbook
import os
import sys

def save_model_results(
    output_file='data.xlsx',
    n_estimators=100,
    min_samples_leaf=35,
    random_seed=42,
    min_sleep_cycle='12*30(6h)',
    new_window_feature_time='无',
    train_sampling_freq='12*5(5分钟)',
    train_set_ratio=0.7,
    test_content='',
    score=0.0,
    model_size='0 MB',
    prediction_rate=0.0,
    prediction_days=0,
    onset_count=0,
    wakeup_count=0,
    match_counts='',
    match_rates='',
    mean_diffs='',
    std_diffs='',
    onset_match_counts='',
    wakeup_match_counts='',
    onset_match_rates='',
    wakeup_match_rates='',
    onset_mean_diffs='',
    wakeup_mean_diffs='',
    onset_std_diffs='',
    wakeup_std_diffs='',
    performance_results=None
):
    """
    保存模型结果到Excel文件
    
    参数:
        output_file: 输出Excel文件名
        n_estimators: 决策树数量
        min_samples_leaf: 叶子节点最小样本数
        random_seed: 随机种子
        min_sleep_cycle: 最小睡眠周期
        new_window_feature_time: 新窗口特征时间
        train_sampling_freq: 训练数据采样频率
        train_set_ratio: 训练集比例
        test_content: 测试内容
        score: 模型得分
        model_size: 模型大小
        prediction_rate: 预测率
        prediction_days: 预测天数
        onset_count: onset事件数量
        wakeup_count: wakeup事件数量
        match_counts: 匹配数
        match_rates: 匹配率
        mean_diffs: 平均差
        std_diffs: 标准差
        onset_match_counts: onset匹配数
        wakeup_match_counts: wakeup匹配数
        onset_match_rates: onset匹配率
        wakeup_match_rates: wakeup匹配率
        onset_mean_diffs: onset平均差
        wakeup_mean_diffs: wakeup平均差
        onset_std_diffs: onset标准差
        wakeup_std_diffs: wakeup标准差
        performance_results: 性能结果字典（包含predicted_onset_count和predicted_wakeup_count）
    
    返回:
        int: 保存的行号
    """
    try:
        # 如果提供了performance_results，则从其中提取数据
        if performance_results is not None:
            try:
                # 提取模型信息（模型大小和预测事件数）
                if 'model_info' in performance_results:
                    model_info = performance_results['model_info']
                    model_size_mb = model_info.get('model_size_mb', 0)
                    model_size = f"{model_size_mb:.2f} MB"
                    onset_count = model_info.get('predicted_onset_count', 0)
                    wakeup_count = model_info.get('predicted_wakeup_count', 0)
                
                # 提取预测统计信息
                if 'prediction_stats' in performance_results:
                    prediction_stats = performance_results['prediction_stats']
                    prediction_rate = prediction_stats.get('prediction_rate', 0.0)
                    prediction_days = prediction_stats.get('total_predicted_days', 0)
                
                # 提取匹配时间统计信息
                if 'match_time_stats' in performance_results:
                    match_time_stats = performance_results['match_time_stats']
                    # 按匹配时间排序
                    sorted_match_times = sorted(match_time_stats.keys())
                    
                    # 提取各种统计数据
                    match_counts_list = []
                    match_rates_list = []
                    mean_diffs_list = []
                    std_diffs_list = []
                    onset_match_counts_list = []
                    wakeup_match_counts_list = []
                    onset_match_rates_list = []
                    wakeup_match_rates_list = []
                    onset_mean_diffs_list = []
                    wakeup_mean_diffs_list = []
                    onset_std_diffs_list = []
                    wakeup_std_diffs_list = []
                    
                    for match_time in sorted_match_times:
                        stats = match_time_stats[match_time]
                        
                        # 总统计
                        total_stats = stats.get('total', {})
                        match_counts_list.append(str(total_stats.get('matches', 0)))
                        match_rates_list.append(f"{total_stats.get('match_rate', 0.0):.4f}")
                        mean_diffs_list.append(f"{total_stats.get('avg_time_diff', 0.0):.2f}")
                        std_diffs_list.append(f"{total_stats.get('std_time_diff', 0.0):.2f}")
                        
                        # Onset统计
                        onset_stats = stats.get('onset', {})
                        onset_match_counts_list.append(str(onset_stats.get('matches', 0)))
                        onset_match_rates_list.append(f"{onset_stats.get('match_rate', 0.0):.4f}")
                        onset_mean_diffs_list.append(f"{onset_stats.get('avg_time_diff', 0.0):.2f}")
                        onset_std_diffs_list.append(f"{onset_stats.get('std_time_diff', 0.0):.2f}")
                        
                        # Wakeup统计
                        wakeup_stats = stats.get('wakeup', {})
                        wakeup_match_counts_list.append(str(wakeup_stats.get('matches', 0)))
                        wakeup_match_rates_list.append(f"{wakeup_stats.get('match_rate', 0.0):.4f}")
                        wakeup_mean_diffs_list.append(f"{wakeup_stats.get('avg_time_diff', 0.0):.2f}")
                        wakeup_std_diffs_list.append(f"{wakeup_stats.get('std_time_diff', 0.0):.2f}")
                    
                    # 转换为字符串，用'/'分隔
                    match_counts = '/'.join(match_counts_list)
                    match_rates = '/'.join(match_rates_list)
                    mean_diffs = '/'.join(mean_diffs_list)
                    std_diffs = '/'.join(std_diffs_list)
                    onset_match_counts = '/'.join(onset_match_counts_list)
                    wakeup_match_counts = '/'.join(wakeup_match_counts_list)
                    onset_match_rates = '/'.join(onset_match_rates_list)
                    wakeup_match_rates = '/'.join(wakeup_match_rates_list)
                    onset_mean_diffs = '/'.join(onset_mean_diffs_list)
                    wakeup_mean_diffs = '/'.join(wakeup_mean_diffs_list)
                    onset_std_diffs = '/'.join(onset_std_diffs_list)
                    wakeup_std_diffs = '/'.join(wakeup_std_diffs_list)
            except (KeyError, AttributeError, TypeError) as e:
                print(f"警告: 从performance_results提取数据时出错: {e}")
                print("使用默认值继续保存数据")
        
        # 验证数据类型
        try:
            onset_count = int(onset_count) if onset_count is not None else 0
            wakeup_count = int(wakeup_count) if wakeup_count is not None else 0
            prediction_rate = float(prediction_rate) if prediction_rate is not None else 0.0
            prediction_days = int(prediction_days) if prediction_days is not None else 0
        except (ValueError, TypeError) as e:
            print(f"警告: 数据类型转换失败: {e}")
            onset_count = 0
            wakeup_count = 0
            prediction_rate = 0.0
            prediction_days = 0
        
        data = [
            test_content,
            n_estimators,
            min_samples_leaf,
            random_seed,
            min_sleep_cycle,
            new_window_feature_time,
            train_sampling_freq,
            train_set_ratio,
            score,
            model_size,
            prediction_rate,
            prediction_days,
            onset_count,
            wakeup_count,
            match_counts,
            match_rates,
            mean_diffs,
            std_diffs,
            onset_match_counts,
            wakeup_match_counts,
            onset_match_rates,
            wakeup_match_rates,
            onset_mean_diffs,
            wakeup_mean_diffs,
            onset_std_diffs,
            wakeup_std_diffs
        ]
        
        file_path = f'/home/zhuangzhuohan/sleep_data/{output_file}'
        
        # 检查文件权限
        try:
            if os.path.exists(file_path):
                if not os.access(file_path, os.R_OK):
                    raise PermissionError(f"无法读取文件: {file_path}")
                if not os.access(file_path, os.W_OK):
                    raise PermissionError(f"无法写入文件: {file_path}")
        except PermissionError as e:
            print(f"错误: {e}")
            return None
        
        # 加载或创建Excel文件
        try:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                next_row = ws.max_row + 1
            else:
                from create_data_xlsx import create_data_xlsx
                create_data_xlsx(output_file)
                wb = load_workbook(file_path)
                ws = wb.active
                next_row = 2
        except Exception as e:
            print(f"错误: 加载或创建Excel文件失败: {e}")
            return None
        
        # 写入数据
        try:
            for col_idx, value in enumerate(data, start=1):
                ws.cell(row=next_row, column=col_idx, value=value)
        except Exception as e:
            print(f"错误: 写入数据到Excel失败: {e}")
            return None
        
        # 保存文件
        try:
            wb.save(file_path)
            print(f"数据已保存到：{file_path}")
            print(f"行号：{next_row}")
            return next_row
        except Exception as e:
            print(f"错误: 保存Excel文件失败: {e}")
            return None
    
    except Exception as e:
        print(f"错误: 保存模型结果时发生未知错误: {e}")
        import traceback
        traceback.print_exc()
        return None
